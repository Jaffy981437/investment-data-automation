"""把最新收盤價回填到大俠武林的 Excel 表(存股紀錄表 / 市值平衡表)。

⚠️ 重要風險
這兩個 xlsx 內含繪圖物件(xl/drawings/*.xml)。openpyxl 儲存時會**重寫整個
檔案**,繪圖物件、部分格式與圖表會遺失,而且無法還原。因此:

  - 預設為預演模式(dry-run),只印出「會改哪些格子、改成什麼」,不動檔案。
  - 只有明確加上 --apply 才會寫入,而且一定先備份到 backups/。

欄位位置不寫死,靠標題列的「商品代號」與「即時股價」自動定位;
表格改版時腳本會直接報錯,而不是默默寫錯欄位。

用法:
    py scripts/update_excel.py                # 預演,安全
    py scripts/update_excel.py --apply        # 實際寫入(會先備份)
"""

from __future__ import annotations

import argparse
import csv
import shutil
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

BASE = Path(__file__).resolve().parent.parent
PRICES = BASE / "data" / "prices" / "latest.csv"
BACKUP_DIR = BASE / "backups"

TARGETS = [
    BASE / "大俠武林" / "大俠客－存股紀錄表(課程訂閱)2024.05.xlsx",
    BASE / "大俠武林" / "大俠客－市值平衡表(課程訂閱)2024.05.xlsx",
]

CODE_HEADER = "商品代號"
PRICE_HEADER = "即時股價"
HEADER_SEARCH_ROWS = 6


def load_prices() -> dict[str, dict]:
    if not PRICES.exists():
        raise SystemExit(f"[x] 找不到 {PRICES.relative_to(BASE)},請先執行 py scripts/fetch_prices.py")
    with PRICES.open(encoding="utf-8-sig", newline="") as f:
        return {r["code"]: r for r in csv.DictReader(f)}


def locate_headers(ws) -> tuple[int, int, int]:
    """回傳 (標題列號, 代號欄號, 股價欄號)。找不到就丟 LookupError。"""
    for row in ws.iter_rows(min_row=1, max_row=HEADER_SEARCH_ROWS):
        code_col = price_col = None
        for cell in row:
            text = str(cell.value).strip() if cell.value is not None else ""
            if text == CODE_HEADER:
                code_col = cell.column
            elif text == PRICE_HEADER:
                price_col = cell.column
        if code_col and price_col:
            return row[0].row, code_col, price_col
    raise LookupError(f"前 {HEADER_SEARCH_ROWS} 列找不到「{CODE_HEADER}」與「{PRICE_HEADER}」標題")


def plan_updates(ws, prices: dict[str, dict]) -> tuple[list[tuple], list[str]]:
    header_row, code_col, price_col = locate_headers(ws)
    updates, misses = [], []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        raw = ws.cell(row=row_idx, column=code_col).value
        if raw is None or str(raw).strip() == "":
            continue
        # Excel 常把 0056 存成數字 56,補回前導零
        code = str(raw).strip()
        if code.replace(".", "").isdigit() and "." not in code:
            code = code.zfill(4)
        quote = prices.get(code)
        if not quote or not quote.get("close"):
            misses.append(code)
            continue
        cell = ws.cell(row=row_idx, column=price_col)
        new_value = float(quote["close"])
        updates.append((cell.coordinate, cell.value, new_value, code,
                        quote.get("name", ""), quote.get("date", "")))
    return updates, misses


def main() -> int:
    ap = argparse.ArgumentParser(description="回填收盤價到大俠武林 Excel")
    ap.add_argument("--apply", action="store_true",
                    help="實際寫入檔案(會先備份)。不加此參數只做預演。")
    args = ap.parse_args()

    prices = load_prices()
    print(f"報價來源:{PRICES.relative_to(BASE)}({len(prices):,} 檔)\n")

    if not args.apply:
        print("=== 預演模式,不會修改任何檔案 ===\n")
    else:
        print("=== 寫入模式 ===")
        print("⚠️  openpyxl 會重寫整個檔案,原檔的繪圖物件與部分格式將永久遺失。\n")

    total = 0
    formula_hits = 0
    for path in TARGETS:
        print("=" * 72)
        print(f"# {path.name}")
        if not path.exists():
            print("  [x] 檔案不存在,略過")
            continue

        wb = load_workbook(path)  # 不用 data_only,保留公式
        changed_any = False
        for ws in wb.worksheets:
            try:
                updates, misses = plan_updates(ws, prices)
            except LookupError:
                continue  # 這張工作表沒有目標欄位(例如「商品資料表」)

            print(f"\n  [{ws.title}] 可更新 {len(updates)} 檔"
                  + (f",查無報價 {len(misses)} 檔:{', '.join(misses[:5])}" if misses else ""))
            for coord, old, new, code, name, date in updates:
                if isinstance(old, str) and ("GOOGLEFINANCE" in old or "DUMMYFUNCTION" in old):
                    formula_hits += 1
                    old_txt = "GOOGLEFINANCE 公式"
                else:
                    old_txt = f"{old:g}" if isinstance(old, (int, float)) else str(old)
                print(f"     {coord}  {code} {name:<14} {old_txt:>18} → {new:<10g} ({date})")
                if args.apply:
                    ws[coord] = new
                    changed_any = True
            total += len(updates)

        if args.apply and changed_any:
            BACKUP_DIR.mkdir(parents=True, exist_ok=True)
            stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup = BACKUP_DIR / f"{path.stem}_{stamp}{path.suffix}"
            shutil.copy2(path, backup)
            wb.save(path)
            print(f"\n  [v] 已備份:{backup.relative_to(BASE)}")
            print(f"  [v] 已寫入:{path.name}")
        wb.close()

    print("\n" + "=" * 72)
    print(f"合計可更新 {total} 檔")

    if formula_hits:
        print(f"\n⚠️  其中 {formula_hits} 格原本是 GOOGLEFINANCE 公式。")
        print("    這些表原本是 Google Sheets,在 Google 試算表裡股價本來就會自動更新;")
        print("    下載成 xlsx 後公式才失效。回填會把公式換成寫死的數字。")
        print("    → 若你主要在 Google 試算表使用這些表,不要回填,讓 GOOGLEFINANCE 自己跑。")
        print("    → 只有需要離線 xlsx 版本時,回填才有意義。")

    if not args.apply:
        print("\n確認無誤後執行:py scripts/update_excel.py --apply")
    return 0


if __name__ == "__main__":
    sys.exit(main())
